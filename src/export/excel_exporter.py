"""
Excel Export Module

This module handles exporting processed Korean vocabulary and grammar to Excel.
"""

import os
import logging
from typing import List, Dict, Any
from pathlib import Path
import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


logger = logging.getLogger(__name__)


class ExcelExporter:
    """Class to export data to Excel."""
    
    def __init__(self, output_path=None):
        """
        Initialize the Excel exporter.
        
        Args:
            output_path: Path to the output Excel file
        """
        self.output_path = output_path or f"korean_vocab_grammar_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    def format_vocabulary_data(self, vocabulary_results: List[Dict]) -> pd.DataFrame:
        """
        Format vocabulary results for Excel export.
        
        Args:
            vocabulary_results: List of dictionaries with vocabulary analysis
            
        Returns:
            Pandas DataFrame with formatted data
        """
        data = []
        
        for item in vocabulary_results:
            data.append({
                'Word': item['item'],
                'Analysis': item['analysis'],
                'Model': item['model'],
                'Has Error': 'Yes' if item.get('error', False) else 'No'
            })
        
        return pd.DataFrame(data)
    
    def format_grammar_data(self, grammar_results: List[Dict]) -> pd.DataFrame:
        """
        Format grammar results for Excel export.
        
        Args:
            grammar_results: List of dictionaries with grammar analysis
            
        Returns:
            Pandas DataFrame with formatted data
        """
        data = []
        
        for item in grammar_results:
            # Extract pattern and example from the item
            original_item = item['item']
            
            # Try to extract pattern and example from the formatted text
            try:
                parts = original_item.split('\n')
                pattern = parts[0].replace('문법: ', '')
                example = parts[1].replace('예문: ', '')
            except:
                pattern = original_item
                example = ""
            
            data.append({
                'Grammar Pattern': pattern,
                'Example': example, 
                'Analysis': item['analysis'],
                'Model': item['model'],
                'Has Error': 'Yes' if item.get('error', False) else 'No'
            })
        
        return pd.DataFrame(data)
    
    def export_to_excel(self, results: Dict) -> str:
        """
        Export processed data to Excel.
        
        Args:
            results: Dictionary with vocabulary and grammar results
            
        Returns:
            Path to the exported Excel file
        """
        logger.info(f"Exporting data to Excel: {self.output_path}")
        
        # Create DataFrames for vocabulary and grammar
        vocabulary_df = self.format_vocabulary_data(results['vocabulary_results'])
        grammar_df = self.format_grammar_data(results['grammar_results'])
        
        # Create a Pandas Excel writer
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            # Write vocabulary data
            vocabulary_df.to_excel(writer, sheet_name='Vocabulary', index=False)
            
            # Write grammar data
            grammar_df.to_excel(writer, sheet_name='Grammar', index=False)
            
            # Auto-adjust column widths for both sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                for idx, col in enumerate(worksheet.columns, 1):
                    max_length = 0
                    column = get_column_letter(idx)
                    
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = min(len(str(cell.value)), 100)  # Limit width to 100 characters
                        except:
                            pass
                    
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column].width = adjusted_width
        
        logger.info(f"Data exported successfully to {self.output_path}")
        return self.output_path
    
    def _apply_formatting(self, workbook: Workbook) -> None:
        """
        Apply formatting to the workbook.
        
        Args:
            workbook: Excel workbook to format
        """
        # Define styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        centered_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Apply formatting to each sheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Format headers
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered_alignment
                cell.border = border
            
            # Adjust row heights
            for i, row in enumerate(worksheet.rows, 1):
                worksheet.row_dimensions[i].height = 20
    
    def apply_formatting(self) -> None:
        """
        Apply formatting to the exported Excel file.
        """
        try:
            from openpyxl import load_workbook
            
            # Load the existing workbook
            workbook = load_workbook(self.output_path)
            
            # Apply formatting
            self._apply_formatting(workbook)
            
            # Save the workbook
            workbook.save(self.output_path)
            logger.info("Applied formatting to Excel file")
            
        except Exception as e:
            logger.warning(f"Failed to apply formatting: {e}")


def export_to_excel(results: Dict, output_path: str = None) -> str:
    """
    Convenience function to export data to Excel.
    
    Args:
        results: Dictionary with vocabulary and grammar results
        output_path: Path to the output Excel file
        
    Returns:
        Path to the exported Excel file
    """
    exporter = ExcelExporter(output_path)
    path = exporter.export_to_excel(results)
    exporter.apply_formatting()
    return path 